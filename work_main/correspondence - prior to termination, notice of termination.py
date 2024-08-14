import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os

# Ścieżka do pliku Excel z danymi klientów
excel_file = "ko.xlsx"

# Ścieżka do folderu z załącznikami
attachments_folder = r""

# Dane do serwera SMTP
smtp_server = "smtp-mail.outlook.com"
smtp_port = 587  # Port serwera SMTP
smtp_username = ""
smtp_password = ""

# Odczyt danych z pliku Excel
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=4, values_only=True):
    message_content = row[1]  # Treść wiadomości z kolumny B
    recipient_email = row[2]  # Adres e-mail z kolumny C
    subject = row[0]  # Temat wiadomości z kolumny A
    document_number = row[3]  # Numer dokumentu z kolumny D

    # Tworzenie wiadomości e-mail
    message = MIMEMultipart()
    message['From'] = smtp_username
    message['To'] = recipient_email
    message['Subject'] = subject


# Dodanie treści wiadomości
    message_content = 'Dzień dobry,' + "</br>" + "</br>" + message_content + "</br>" + "</br>" \
    'Informujemy, że od 17.05.2024 r., spółka Digital Care sp. z o.o. działa pod nazwą „bolttech Poland sp. z o.o.”. Zmiana nazwy Spółki w żaden sposób nie wpływa na obowiązki Digital Care sp. z o.o. (obecnie bolttech Poland) wobec jej klientów i kontrahentów.'"</br>" + "</br>"\
    'Masz pytania lub potrzebujesz pomocy? Skontaktuj się z nami pisząc na adres najem@digitalcaregroup.com lub dzwoniąc pod numer +48 (22) 829-99-09'\
    + "</br>" + "</br>" + 'Pozdrawiamy,' + "</br>" 'Zespół Digital Care'\
    + "</br>" + "</br>"\
    + "<html>\
                        <body>\
                            <table style=\"background: #FFF; margin-left: auto; margin-right: auto; max-width: 600px;\" width=\"600\">\
                            <tbody>\
                                <tr>\
                                    <td style=\"line-height: 13px; color: #686868; font-size: 10px; padding: 20px 50px; text-align: left\" colspan=\"2\">\
                                    <p style=\"border-bottom: 1px solid #E3E3E3; padding: 20px 0px; margin: 0;\">\
                                    <img alt=\"DC\" height=\"17\" src=\"https://mail.digitalcaregroup.com/mail/imad/shared/logo%20dc.png\" width=\"90\" style=\"vertical-align: middle, inline;\" />\
                                        <img alt=\"DC\" height=\"45\" src=\"https://mail.digitalcaregroup.com/mail/dc/sig/medium_04.png\" width=\"31\" style=\"vertical-align: middle; padding-left: 350px;\" />\
                                            </p>\
                                                </td>\
                                                    </tr>\
                                                        <tr>\
                                                            <td style=\"color: #686868; font-size: 10px; line-height: 13px; font-weight: normal; padding: 13px 0 13px 40px; width: 200px; font-family: Arial, Helvetica, sans-serif;\">bolttech Poland Sp. z o.o. <br />\
                                                            (poprzednia nazwa Digital Care sp. z o.o.) <br />\
                                                            ul. Marszałkowska 126/134 <br />\
                                                                00-008 Warszawa | Polska</td>\
                                                                    <td style=\"color: #686868; font-size: 10px; line-height: 13px; font-weight: normal; font-family: Arial, Helvetica, sans-serif;\">www.digitalcaregroup.pl</td>\
                                                                    </tr>\
                                                                        <tr>\
                                                                            <td style=\"line-height: 13px; font-size: 10px; padding: 0px 40px 25px; text-align: left\" colspan=\"2\">\
                                                                                <p style=\"color: #686868; font-family: Arial, Helvetica, sans-serif; line-height: 14px; font-size: 10px; font-weight: 400; margin: 0; padding: 4px 0 15px;\">NIP: 5342487752 | Numer rejestrowy BDO: 000527426 | Kapitał zakładowy: 9 888 300 zł, wpłacony w całości<br />KRS: 0000431665, Sąd Rejonowy dla m. st. Warszawy w Warszawie, XII Wydział Gospodarczy Krajowego Rejestru Sądowego</p>\
                                                                                </td>\
                                                                                    </tr>\
                                                                                        </tbody>\
                                                                                            </table>\
                                                                                                </body>\
                                                                                                    </html>" # Dodanie kodu HTML elementu do treści wiadomości
    message.attach(MIMEText(message_content, 'html'))

    # Dodanie załącznika
    attachment_path = os.path.join(attachments_folder, f"{document_number}.pdf")
    with open(attachment_path, 'rb') as attachment:
        attachment_mime = MIMEApplication(attachment.read(), _subtype="pdf")
        attachment_mime.add_header('content-disposition', 'attachment', filename=f"{document_number}.pdf")
        message.attach(attachment_mime)

    # Inicjalizacja serwera SMTP
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()  # Rozpoczęcie szyfrowanego połączenia
        server.login(smtp_username, smtp_password)  # Logowanie do konta SMTP
        server.sendmail(smtp_username, recipient_email, message.as_string())

print("Wysłano maile do wszystkich klientów.")
